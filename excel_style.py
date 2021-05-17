import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

import downloading
import opdownloading
range_cell = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11']
range_cell2 = ['G2', 'G3', 'G4', 'G5', 'G6', 'G7', 'G8', 'G9', 'G10',
               'K2', 'K3', 'K4', 'K5', 'K6', 'K7', 'K8', 'K9', 'K10',
               'O2', 'O3', 'O4', 'O5', 'O6', 'O7', 'O8', 'O9', 'O10']
range_cell3 = ['E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'E13', 'E14',
               'I2', 'I3', 'I4', 'I5', 'I6', 'I7', 'I8', 'I9', 'I10', 'I11', 'I12', 'I13', 'I14',
               'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9', 'M10', 'M11', 'M12', 'M13', 'M14']

def style1(file):
    wb = load_workbook(file)
    ws = wb.active

    for col in ws.columns:
        for c in col:
            c.border = Border(left=Side(border_style='thin',
                                        color='FF000000'),
                              right=Side(border_style='thin',
                                         color='FF000000'),
                              top=Side(border_style='thin',
                                       color='FF000000'),
                              bottom=Side(border_style='thin',
                                          color='FF000000'),
                              )
            c.alignment = Alignment(horizontal='center')
            c.font = Font(name='微软雅黑', size=10, )

    row = ws[1]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True)
        c.fill = PatternFill(fgColor='FFD9E1F2', fill_type='solid')

    row = ws[8]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True)
        c.fill = PatternFill(fgColor='FFD9E1F2', fill_type='solid')

    row = ws[13]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True)
        c.fill = PatternFill(fgColor='FFD9E1F2', fill_type='solid')

    row = ws[14]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True)
        c.fill = PatternFill(fgColor='FFD9E1F2', fill_type='solid')

    for cell in range_cell3:
        if ws[cell].value < 0.75:
            ws[cell].fill = PatternFill(fgColor='FF0000', fill_type='solid')

    col = ws['E']
    for c in col:
        c.number_format = '0%'
    col = ws['I']
    for c in col:
        c.number_format = '0%'
    col = ws['M']
    for c in col:
        c.number_format = '0%'

    col = ws['A']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True)
    col = ws['D']
    for c in col:
        c.fill = PatternFill(fgColor='FF70AD47', fill_type='solid')
    col = ws['H']
    for c in col:
        c.fill = PatternFill(fgColor='FF70AD47', fill_type='solid')
    col = ws['L']
    for c in col:
        c.fill = PatternFill(fgColor='FF70AD47', fill_type='solid')


    st.sidebar.markdown(opdownloading.get_table_download_link(wb, '组织区域满编率', '组织区域满编率'), unsafe_allow_html=True)

def style2(file):
    wb=load_workbook(file)
    ws = wb.active
    ws.insert_cols(idx=1)
    range_cell = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11']
    range_cell2 = ['G2', 'G3', 'G4', 'G5', 'G6', 'G7', 'G8', 'G9', 'G10',
                   'K2', 'K3', 'K4', 'K5', 'K6', 'K7', 'K8', 'K9', 'K10',
                   'O2', 'O3', 'O4', 'O5', 'O6', 'O7', 'O8', 'O9', 'O10']
    range_cell3 = ['E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'E13', 'E14',
                   'I2', 'I3', 'I4', 'I5', 'I6', 'I7', 'I8', 'I9', 'I10', 'I11', 'I12', 'I13', 'I14',
                   'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9', 'M10', 'M11', 'M12', 'M13', 'M14']

    for c in range_cell:
        ws[c] = '上海三区'

    ws['A1'] = '区域'
    ws['C11'] = '上海三区'
    ws.merge_cells('A11:C11')
    ws.column_dimensions['B'].width = 25.0

    for col in ws.columns:
        for c in col:
            c.border = Border(left=Side(border_style='thin',
                                        color='FF000000'),
                              right=Side(border_style='thin',
                                         color='FF000000'),
                              top=Side(border_style='thin',
                                       color='FF000000'),
                              bottom=Side(border_style='thin',
                                          color='FF000000'),
                              )
            c.alignment = Alignment(horizontal='center')
            c.font = Font(name='微软雅黑', size=10, )

    row = ws[1]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')

    row = ws[11]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF')
        c.fill = PatternFill(fgColor='FF8497B0', fill_type='solid')

    for cell in range_cell2:
        if ws[cell].value < 0.75:
            ws[cell].fill = PatternFill(fgColor='FF0000', fill_type='solid')

    col = ws['G']
    for c in col:
        c.number_format = '0%'
    col = ws['K']
    for c in col:
        c.number_format = '0%'
    col = ws['O']
    for c in col:
        c.number_format = '0%'


    st.sidebar.markdown(opdownloading.get_table_download_link(wb, '校区满编率', '校区满编率'), unsafe_allow_html=True)

def style3(file):
    wb=load_workbook(file)
    ws = wb.active
    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:I1')
    ws.merge_cells('J1:Q1')
    ws.merge_cells('R1:Y1')
    ws['A1'] = '大区/区域'
    ws['B1'] = '入职'
    ws['J1'] = '离职'
    ws['R1'] = '净增长'

    for col in ws.columns:
        for c in col:
            c.border = Border(left=Side(border_style='thin',
                                        color='FF000000'),
                              right=Side(border_style='thin',
                                         color='FF000000'),
                              top=Side(border_style='thin',
                                       color='FF000000'),
                              bottom=Side(border_style='thin',
                                          color='FF000000'),
                              )
            c.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            c.font = Font(name='微软雅黑', size=12, )

    ws.column_dimensions['A'].width = 10.0
    column = 2
    while column < 26:
        i = get_column_letter(column)
        ws.column_dimensions[i].width = 5
        column += 1

    row = ws[1]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    row = ws[2]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    row = ws[9]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    row = ws[14]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    row = ws[15]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    col = ws['A']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
    col = ws['I']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    col = ws['Q']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')
    col = ws['Y']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, size=12, )
        c.fill = PatternFill(fgColor='FFDDEBF7', fill_type='solid')



    st.sidebar.markdown(opdownloading.get_table_download_link(wb, '组织区域入职离职数据', '组织区域入职离职数据'), unsafe_allow_html=True)

def style4(file):
    wb=load_workbook(file)
    ws = wb.active
    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:H1')
    ws.merge_cells('I1:O1')
    ws.merge_cells('P1:V1')
    ws['A1'] = '校区'
    ws['B1'] = '入职'
    ws['I1'] = '离职'
    ws['P1'] = '净增长'

    for col in ws.columns:
        for c in col:
            c.border = Border(left=Side(border_style='thin',
                                        color='FF000000'),
                              right=Side(border_style='thin',
                                         color='FF000000'),
                              top=Side(border_style='thin',
                                       color='FF000000'),
                              bottom=Side(border_style='thin',
                                          color='FF000000'),
                              )
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name='微软雅黑', size=10, )

    ws.column_dimensions['A'].width = 25.0
    row = ws[1]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')
    row = ws[2]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')
    row = ws[12]
    for c in row:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')
    col = ws['H']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')
    col = ws['O']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')
    col = ws['V']
    for c in col:
        c.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10, )
        c.fill = PatternFill(fgColor='FF5B9BD5', fill_type='solid')

    st.sidebar.markdown(opdownloading.get_table_download_link(wb, '校区入职离职数据', '校区入职离职数据'), unsafe_allow_html=True)